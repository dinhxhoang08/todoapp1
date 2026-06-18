import csv
from io import StringIO, BytesIO
from typing import List
from sqlalchemy.orm import Session
from app.repositories.task_repository import TaskRepository
from app.models.task import Task
from openpyxl import Workbook
from openpyxl.styles import Font


class ExportService:
    def __init__(self, db: Session):
        self.db = db
        self.task_repo = TaskRepository(db)

    def _get_tasks(self, user_id: int) -> List[Task]:
        return self.task_repo.list_by_user(user_id, limit=10000)

    def export_tasks_csv(self, user_id: int) -> StringIO:
        tasks = self._get_tasks(user_id)
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Title", "Description", "Status", "Priority", "Due Date", "Completed At", "Created At"])
        for t in tasks:
            writer.writerow([
                t.id, t.title, t.description or "",
                t.status.value if t.status else "",
                t.priority.value if t.priority else "",
                t.due_date.isoformat() if t.due_date else "",
                t.completed_at.isoformat() if t.completed_at else "",
                t.created_at.isoformat() if t.created_at else "",
            ])
        output.seek(0)
        return output

    def export_tasks_excel(self, user_id: int) -> BytesIO:
        tasks = self._get_tasks(user_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        headers = ["ID", "Title", "Description", "Status", "Priority", "Due Date", "Completed At", "Created At"]
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
        for row, t in enumerate(tasks, 2):
            ws.cell(row=row, column=1, value=t.id)
            ws.cell(row=row, column=2, value=t.title)
            ws.cell(row=row, column=3, value=t.description or "")
            ws.cell(row=row, column=4, value=t.status.value if t.status else "")
            ws.cell(row=row, column=5, value=t.priority.value if t.priority else "")
            ws.cell(row=row, column=6, value=t.due_date.isoformat() if t.due_date else "")
            ws.cell(row=row, column=7, value=t.completed_at.isoformat() if t.completed_at else "")
            ws.cell(row=row, column=8, value=t.created_at.isoformat() if t.created_at else "")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
